#from openpyxl import workbook
import gi

gi.require_version("Gtk", "3.0")
from gi.repository import Gtk
from recipes import burger
from recipes import pizza
from recipes import wok
from openpyxl import Workbook

wb = Workbook()

ws1 = wb.active
ws1.title = "range names"

dest_filename = 'empty_book.xlsx'


class ListBoxRowWithData(Gtk.ListBoxRow):
    def __init__(self, data):
        super().__init__()
        self.data = data
        self.add(Gtk.Label(label=data))


class MyWindow(Gtk.Window):
    def __init__(self):
        super().__init__(title="Кухня")

        self.box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=10)
        self.add(self.box)

        self.burger_button = Gtk.Button(label="Бургер")
        self.burger_button.connect("clicked", self.burger_button_clicked)
        self.box.pack_start(self.burger_button, True, True, 0)

        self.pizza_button = Gtk.Button(label="Пицца")
        self.pizza_button.connect("clicked", self.pizza_button_clicked)
        self.box.pack_start(self.pizza_button, True, True, 0)

        self.wok_button = Gtk.Button(label="Вок")
        self.wok_button.connect("clicked", self.wok_button_clicked)
        self.box.pack_start(self.wok_button, True, True, 0)

        self.energy = 0
        self.cost = 0
        self.row = 1
        lok = "Выберите закуску"
        self.lbl = Gtk.Label(
            label=lok
        )
        self.box.pack_start(self.lbl, True, True, 0)

        self.save_button = Gtk.Button(label="Сохранить")
        self.save_button.connect("clicked", self.save_button_clicked)
        self.box.pack_start(self.save_button, True, True, 0)

        spis = "Список ингредиентов"
        self.list = Gtk.Label(label=spis)
        self.box.pack_start(self.list, True, True, 0)

        self.ingredients_list = Gtk.ListBox()
        self.box.pack_start(self.ingredients_list, True, True, 0)

    def burger_button_clicked(self, widget):
        self.energy += burger.e()
        self.cost += burger.c()
        label = (burger.res+"\nВы получили "+str(round(self.energy, 2)) + " энергии за " + str(round(self.cost, 2)) + " рублей!")
        lok = Gtk.Label.set_text(self.lbl, label)
        self.ingredients_list.destroy()
        self.ingredients_list = Gtk.ListBox()
        self.box.pack_start(self.ingredients_list, True, True, 0)
        for key in burger.recipe:
            self.ingredients_list.add(ListBoxRowWithData(key+" "+str(burger.recipe[key]*100)+" г"))
            self.ingredients_list.show_all()
            ws1.cell(row=self.row, column=1).value = key+" "+str(burger.recipe[key]*100)+" г"
            self.row += 1
            wb.save(filename=dest_filename)
        self.box.pack_start(lok, True, True, 0)

    def pizza_button_clicked(self, widget):
        self.energy += pizza.e()
        self.cost += pizza.c()
        label = (pizza.res + "\nВы получили " + str(round(self.energy, 2)) + " энергии за " + str(round(self.cost, 2)) + " рублей!")
        lok = Gtk.Label.set_text(self.lbl, label)
        self.ingredients_list.destroy()
        self.ingredients_list = Gtk.ListBox()
        self.box.pack_start(self.ingredients_list, True, True, 0)
        for key in pizza.recipe:
            self.ingredients_list.add(ListBoxRowWithData(key+" "+str(pizza.recipe[key]*100)+" г"))
            self.ingredients_list.show_all()
            ws1.cell(row=self.row, column=1).value = key+" "+str(pizza.recipe[key]*100)+" г"
            self.row += 1
            wb.save(filename=dest_filename)
        self.box.pack_start(lok, True, True, 0)

    def wok_button_clicked(self, widget):
        self.energy += wok.e()
        self.cost += wok.c()
        label = (wok.res + "\nВы получили " + str(round(self.energy, 2)) + " энергии за " + str(round(self.cost, 2)) + " рублей!")
        lok = Gtk.Label.set_text(self.lbl, label)
        self.ingredients_list.destroy()
        self.ingredients_list = Gtk.ListBox()
        self.box.pack_start(self.ingredients_list, True, True, 0)
        for key in wok.recipe:
            self.ingredients_list.add(ListBoxRowWithData(key+" "+str(wok.recipe[key]*100)+" г"))
            self.ingredients_list.show_all()
            ws1.cell(row=self.row, column=1).value = key+" "+str(wok.recipe[key]*100)+" г"
            self.row += 1
            wb.save(filename=dest_filename)
        self.box.pack_start(lok, True, True, 0)

    def save_button_clicked(self, widget):
        lok = Gtk.Label.set_text(self.lbl, "Расчёты сохранены!")
        ws1['D1'].value = "Общая стоимость: "+str(round(self.cost, 2))+" рублей"
        ws1['D2'].value = "Общая калорийность: "+str(round(self.energy, 2))+" калорий"
        wb.save(filename=dest_filename)
        self.box.pack_start(lok, True, True, 0)


win = MyWindow()
win.connect("destroy", Gtk.main_quit)
win.show_all()
Gtk.main()
#wb = workbook()
