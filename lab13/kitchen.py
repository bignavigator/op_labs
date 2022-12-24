from openpyxl import Workbook
from guizero import App, Text, PushButton, ListBox

from recipes.burger import Burger
from recipes.pizza import Pizza
from recipes.wok import Wok

energy = 0
cost = 0
row = 1
wb = Workbook()

ws1 = wb.active
ws1.title = "range names"

dest_filename = 'energies_and_costs.xlsx'


def burger_button_clicked():
    r = Burger.burger_recipe
    t = Burger.res
    a = Burger()
    e = a.e()
    c = a.c()
    global energy
    energy += e
    global cost
    cost += c
    label = (t+"\nВы получили "+str(round(energy, 2)) + " энергии за " + str(round(cost, 2)) + " рублей!")
    lok.value = label
    listbox.clear()
    global row
    for key in r:
        listbox.append(key + " " + str(r[key] * 100) + " г")
        ws1.cell(row=row, column=1).value = key + " " + str(r[key] * 100) + " г"
        row += 1


def pizza_button_clicked():
    r = Pizza.pizza_recipe
    t = Pizza.res
    a = Pizza()
    e = a.e()
    c = a.c()
    global energy
    energy += e
    global cost
    cost += c
    label = (t+"\nВы получили "+str(round(energy, 2)) + " энергии за " + str(round(cost, 2)) + " рублей!")
    lok.value = label
    listbox.clear()
    global row
    for key in r:
        listbox.append(key + " " + str(r[key] * 100) + " г")
        ws1.cell(row=row, column=1).value = key + " " + str(r[key] * 100) + " г"
        row += 1


def wok_button_clicked():
    r = Wok.wok_recipe
    t = Wok.res
    a = Wok()
    e = a.e()
    c = a.c()
    global energy
    energy += e
    global cost
    cost += c
    label = (t+"\nВы получили "+str(round(energy, 2)) + " энергии за " + str(round(cost, 2)) + " рублей!")
    lok.value = label
    listbox.clear()
    global row
    for key in r:
        listbox.append(key + " " + str(r[key] * 100) + " г")
        ws1.cell(row=row, column=1).value = key + " " + str(r[key] * 100) + " г"
        row += 1


def save_button_clicked():
    label = "Расчёты сохранены!"
    lok.value = label
    ws1['D1'].value = "Общая стоимость: " + str(round(cost, 2)) + " рублей"
    ws1['D2'].value = "Общая калорийность: " + str(round(energy, 2)) + " калорий"
    wb.save('energies_and_costs.xlsx')


app = App(title="Кухня")
lok = Text(app, text="Выберите рецепты")
listbox = ListBox(app, items=[], width=150, height=250)
burger_button = PushButton(app, text="Бургер", command=burger_button_clicked)
pizza_button = PushButton(app, text="Пицца", command=pizza_button_clicked)
wok_button = PushButton(app, text="Вок", command=wok_button_clicked)
save_button = PushButton(app, text="Сохранить", command=save_button_clicked)
app.display()
